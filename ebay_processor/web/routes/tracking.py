# ebay_processor/web/routes/tracking.py
"""
Tracking File Upload Routes.

This module manages the upload of CSV files from carriers
with tracking numbers and updates the tracking files
generated by the application.
"""

import logging
import os
import shutil
import random
import glob
from datetime import datetime, timezone

import pandas as pd
import openpyxl
from flask import (
    Blueprint,
    render_template,
    request,
    redirect,
    url_for,
    flash,
    session,
    current_app,
)
from werkzeug.utils import secure_filename

from ..decorators import login_required
from ...core.constants import INFO_HEADER_TAG

logger = logging.getLogger(__name__)

# Create the Blueprint for tracking routes.
tracking_bp = Blueprint(
    'tracking',
    __name__,
    template_folder='../../templates',
    static_folder='../../static'
)


@tracking_bp.route('/upload-tracking', methods=['GET'])
@login_required
def upload_tracking_form():
    """Shows the form to upload the tracking file."""
    output_dir = current_app.config['OUTPUT_DIR']
    
    # Search for existing tracking files to show them in the form.
    try:
        tracking_pattern = os.path.join(output_dir, '*Tracking*.xlsx')
        # Use a generator comprehension for efficiency
        tracking_files = (os.path.basename(f) for f in glob.glob(tracking_pattern) if os.path.isfile(f))
        # Sort the final list
        sorted_files = sorted(list(tracking_files))
        
        # Also look for demo CSV files for automatic suggestion
        csv_pattern = os.path.join(output_dir, '*COURIER_UPLOAD_DEMO.csv')
        demo_csv_files = sorted([os.path.basename(f) for f in glob.glob(csv_pattern) if os.path.isfile(f)])
        
        # Find the most recent consolidated CSV for demo suggestion
        consolidated_csv = None
        for csv_file in reversed(demo_csv_files):  # Most recent first due to timestamp
            if 'CONSOLIDATED' in csv_file:
                consolidated_csv = csv_file
                break
                
    except Exception as e:
        logger.error(f"Error listing tracking files: {e}", exc_info=True)
        flash("Could not get the list of existing tracking files.", "error")
        sorted_files = []
        demo_csv_files = []
        consolidated_csv = None
        
    return render_template('upload_tracking.html', 
                         existing_files=sorted_files,
                         demo_csv_files=demo_csv_files,
                         suggested_csv=consolidated_csv)


@tracking_bp.route('/process-tracking-upload', methods=['POST'])
@login_required
def process_tracking_upload():
    """
    Processes the uploaded tracking file and updates the corresponding Excel files.
    """
    # --- 1. Uploaded Tracking File Validation ---
    # Check if using demo CSV selection
    selected_demo_csv = request.form.get('selected_demo_csv', '').strip()
    
    if selected_demo_csv:
        # Using demo CSV file
        output_dir = current_app.config['OUTPUT_DIR']
        tracking_file_path = os.path.join(output_dir, selected_demo_csv)
        
        if not os.path.exists(tracking_file_path):
            flash(f'The selected demo file does not exist: {selected_demo_csv}', 'error')
            return redirect(url_for('tracking.upload_tracking_form'))
            
        logger.info(f"Using demo CSV file: {selected_demo_csv}")
        
        # Skip file upload processing, use the existing demo file
        temp_dir = None  # No temp dir needed for demo files
        
    else:
        # Regular file upload
        if 'tracking_file' not in request.files:
            flash('No tracking file found in the request.', 'error')
            return redirect(url_for('tracking.upload_tracking_form'))

        tracking_file = request.files['tracking_file']
        if tracking_file.filename == '':
            flash('No tracking file was selected.', 'error')
            return redirect(url_for('tracking.upload_tracking_form'))

        if not tracking_file.filename.lower().endswith('.csv'):
            flash('The tracking file must be a CSV file.', 'error')
            return redirect(url_for('tracking.upload_tracking_form'))
            
        # --- 2. Create a Temporary Directory for the Operation ---
        temp_dir = os.path.join(current_app.config['OUTPUT_DIR'], f"temp_upload_{datetime.now().strftime('%Y%m%d%H%M%S')}")
        os.makedirs(temp_dir, exist_ok=True)
        tracking_file_path = os.path.join(temp_dir, secure_filename(tracking_file.filename))
        tracking_file.save(tracking_file_path)

    # --- 3. Read and Process the Tracking CSV ---
    try:
        # Try reading with various common encodings.
        tracking_df = None
        for enc in ['utf-8', 'cp1252', 'latin1']:
            try:
                tracking_df = pd.read_csv(tracking_file_path, encoding=enc)
                logger.info(f"Tracking file read with encoding: {enc}")
                break
            except UnicodeDecodeError:
                continue
        
        if tracking_df is None:
            raise ValueError("Could not decode the CSV file. Please ensure it is in UTF-8, CP1252 or Latin-1 format.")

        tracking_df.columns = map(str.lower, tracking_df.columns)
        if 'order number' not in tracking_df.columns or 'consignment number' not in tracking_df.columns:
            raise ValueError("The CSV file must contain 'Order Number' and 'Consignment Number' columns.")

        # Create a map of Barcode -> Tracking Number.
        tracking_map = {
            str(row['order number']).strip(): str(row['consignment number']).strip()
            for _, row in tracking_df.iterrows()
            if pd.notna(row['order number']) and pd.notna(row['consignment number'])
        }
        logger.info(f"Tracking map created with {len(tracking_map)} entries.")

    except (ValueError, Exception) as e:
        flash(f"Error processing tracking CSV file: {e}", 'error')
        if temp_dir and os.path.exists(temp_dir):
            shutil.rmtree(temp_dir)
        return redirect(url_for('tracking.upload_tracking_form'))

    # --- 4. Process Target Excel Files ---
    selected_files = request.form.getlist('selected_files')
    update_all = request.form.get('update_all') == 'on'
    
    # If "update all" is checked, get all available tracking files
    if update_all and not selected_files:
        output_dir = current_app.config['OUTPUT_DIR']
        try:
            tracking_pattern = os.path.join(output_dir, '*Tracking*.xlsx')
            all_tracking_files = [os.path.basename(f) for f in glob.glob(tracking_pattern) if os.path.isfile(f)]
            selected_files = sorted(all_tracking_files)
            logger.info(f"Update all selected - processing {len(selected_files)} files: {selected_files}")
        except Exception as e:
            logger.error(f"Error getting tracking files for update all: {e}", exc_info=True)
    
    if not selected_files:
        flash('No Excel files were selected for update.', 'warning')
        if temp_dir and os.path.exists(temp_dir):
            shutil.rmtree(temp_dir)
        return redirect(url_for('tracking.upload_tracking_form'))

    updated_files_info = []
    total_updates = 0
    output_dir = current_app.config['OUTPUT_DIR']

    try:
        for filename in selected_files:
            safe_filename = secure_filename(filename)
            file_path = os.path.join(output_dir, safe_filename)
            
            if not os.path.isfile(file_path):
                flash(f"File '{safe_filename}' was not found. Skipping.", 'warning')
                continue

            try:
                updates_in_file, new_file_path = _update_excel_file(file_path, tracking_map, output_dir, session.get('user_id', 'user'))
                if updates_in_file > 0:
                    updated_files_info.append({
                        'original_file': safe_filename,
                        'updated_file': os.path.basename(new_file_path),
                        'updates': updates_in_file,
                    })
                    total_updates += updates_in_file
                else:
                    flash(f"No barcode matches found in file '{safe_filename}'.", 'info')
            except Exception as e:
                logger.error(f"Failed to update Excel file '{safe_filename}': {e}", exc_info=True)
                flash(f"An error occurred while processing file '{safe_filename}': {e}", 'error')

        # --- 5. Show Results ---
        flash(f'Process completed. Updated {total_updates} tracking numbers in {len(updated_files_info)} file(s).', 'success')
        return render_template('tracking_upload_results.html', updated_files=updated_files_info, total_updates=total_updates)

    finally:
        # --- 6. Temporary Directory Cleanup ---
        if temp_dir and os.path.exists(temp_dir):
            shutil.rmtree(temp_dir)
            logger.info(f"Tracking upload temporary directory deleted: {temp_dir}")


def _update_excel_file(original_path: str, tracking_map: dict, output_dir: str, username: str) -> tuple[int, str]:
    """Helper function to update a single Excel file."""
    wb = openpyxl.load_workbook(original_path)
    ws = wb.active # Assume data is in the first active sheet.

    # Determine the header row.
    header_row_index = 2 if ws.cell(row=1, column=1).value == INFO_HEADER_TAG else 1
    headers = {str(cell.value).lower().strip(): idx for idx, cell in enumerate(ws[header_row_index], 1) if cell.value}
    
    barcode_col = headers.get('our_barcode')
    tracking_col = headers.get('tracking number')

    if not barcode_col or not tracking_col:
        raise ValueError("The Excel file does not contain 'Our_Barcode' or 'Tracking Number' columns.")

    updates = 0
    for row in range(header_row_index + 1, ws.max_row + 1):
        barcode_cell = ws.cell(row=row, column=barcode_col)
        barcode_value = str(barcode_cell.value).strip() if barcode_cell.value else None
        
        if barcode_value in tracking_map:
            tracking_number = tracking_map[barcode_value]
            tracking_cell = ws.cell(row=row, column=tracking_col)
            tracking_cell.value = tracking_number
            tracking_cell.number_format = '@' # Ensure it's saved as text.
            updates += 1
    
    if updates > 0:
        base, ext = os.path.splitext(os.path.basename(original_path))
        timestamp = datetime.now(timezone.utc).strftime('%Y%m%d_%H%M')
        new_filename = f"{base}_updated_{username}_{timestamp}{ext}"
        new_path = os.path.join(output_dir, new_filename)
        wb.save(new_path)
        return updates, new_path
    
    return 0, ""