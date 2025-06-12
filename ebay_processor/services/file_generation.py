# ebay_processor/services/file_generation.py
"""
File Generation Service.

This module is responsible for creating all output files, primarily
Excel spreadsheets (RUN, COURIER_MASTER, Tracking, etc.).
Takes already processed data and formats it according to each file's specifications.
"""
import logging
import os
import shutil
from datetime import datetime
from typing import List, Dict, Any, Optional

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# Import project utilities and constants
from ..core.constants import (
    HIGHLANDS_AND_ISLANDS_POSTCODES,
    NEXT_DAY_SERVICE_NAME,
    STANDARD_SERVICE_NAME,
    INFO_HEADER_TAG,
    RUN_SHEET_TITLE,
    RUN24H_SHEET_TITLE,
    COURIER_MASTER_SHEET_TITLE,
    TRACKING_SHEET_TITLE,
    UNMATCHED_SHEET_TITLE,
)
from ..core.exceptions import FileGenerationError
from ..utils.string_utils import sanitize_for_excel

logger = logging.getLogger(__name__)


# --- Public File Generation Functions ---

def generate_consolidated_run_file(
    orders: List[Dict[str, Any]], output_dir: str, run_date: datetime, config: Dict
) -> Optional[str]:
    """
    Generates a single consolidated RUN file for standard orders.

    Args:
        orders: List of standard order items.
        output_dir: Directory where the file will be saved.
        run_date: Execution date for the filename.
        config: Application configuration dictionary.

    Returns:
        The path to the generated file or None if not generated.
    """
    logger.info(f"Generating consolidated RUN file for {len(orders)} standard items.")
    if not orders:
        return None

    filename = _generate_filename("RUN", None, run_date, consolidated=True)
    rows = [_format_run_row(item) for item in orders]
    
    return _save_excel_file(rows, output_dir, filename, RUN_SHEET_TITLE, config)


def generate_run24h_file(
    orders: List[Dict[str, Any]], output_dir: str, run_date: datetime, config: Dict
) -> Optional[str]:
    """
    Generates a single consolidated RUN24H file for urgent/expedited orders.

    Args:
        orders: List of urgent order items.
        output_dir: Directory where the file will be saved.
        run_date: Execution date for the filename.
        config: Application configuration dictionary.

    Returns:
        The path to the generated file or None if not generated.
    """
    logger.info(f"Generating consolidated RUN24H file for {len(orders)} urgent items.")
    if not orders:
        return None

    filename = _generate_filename("RUN24H", None, run_date, consolidated=True)
    rows = [_format_run_row(item) for item in orders]

    return _save_excel_file(rows, output_dir, filename, RUN24H_SHEET_TITLE, config)


def generate_consolidated_courier_master_file(
    all_orders: List[Dict[str, Any]], output_dir: str, run_date: datetime, config: Dict
) -> Optional[str]:
    """
    Generates a single consolidated COURIER_MASTER file with one row per order.

    Args:
        all_orders: List of all processed items.
        output_dir: Directory where the file will be saved.
        run_date: Execution date for the filename.
        config: Application configuration dictionary.

    Returns:
        The path to the generated file or None if not generated.
    """
    logger.info(f"Generating COURIER_MASTER for a total of {len(all_orders)} items.")
    if not all_orders:
        return None
        
    order_groups = _group_items_by_order_id(all_orders)
    rows = []
    for order_id, items in order_groups.items():
        first_item = items[0]
        rows.append(_format_courier_master_row(first_item, items))

    if rows:
        filename = _generate_filename("COURIER_MASTER", None, run_date, consolidated=True, config=config)
        return _save_excel_file(rows, output_dir, filename, COURIER_MASTER_SHEET_TITLE, config)
    return None


def generate_tracking_files(
    all_orders: List[Dict[str, Any]], output_dir: str, run_date: datetime, config: Dict
) -> List[str]:
    """
    Generates multiple Tracking files: one consolidated and one per store.
    Also generates demo CSV files for tracking upload testing.

    Args:
        all_orders: List of all processed items.
        output_dir: Directory where files will be saved.
        run_date: Execution date for filenames.
        config: Application configuration dictionary.

    Returns:
        A list of paths to all generated files (Excel and CSV).
    """
    logger.info("Starting generation of all Tracking files.")
    generated_paths = []

    # 1. Generate consolidated tracking file.
    consolidated_excel, consolidated_csv = _create_single_tracking_file_with_csv(
        all_orders, output_dir, run_date, config, is_consolidated=True
    )
    if consolidated_excel:
        generated_paths.append(consolidated_excel)
    if consolidated_csv:
        generated_paths.append(consolidated_csv)

    # 2. Generate tracking files by store.
    store_groups = _group_items_by_store_id(all_orders)
    for store_id, items in store_groups.items():
        store_excel, store_csv = _create_single_tracking_file_with_csv(
            items, output_dir, run_date, config, is_consolidated=False, store_id=store_id
        )
        if store_excel:
            generated_paths.append(store_excel)
        if store_csv:
            generated_paths.append(store_csv)
            
    return generated_paths


def generate_unmatched_items_file(
    unmatched_items: List[Dict[str, Any]], output_dir: str, run_date: datetime, config: Dict
) -> Optional[str]:
    """
    Generates an Excel file with all items that couldn't be matched.

    Args:
        unmatched_items: List of dictionaries of unmatched items.
        output_dir: Directory where the file will be saved.
        run_date: Execution date for the filename.
        config: Application configuration dictionary.

    Returns:
        The path to the generated file or None if there were no unmatched items.
    """
    logger.info(f"Generating unmatched items file for {len(unmatched_items)} items.")
    if not unmatched_items:
        return None
    
    filename = f"unmatched_items_{run_date.strftime('%Y%m%d_%H%M%S')}.xlsx"
    return _save_excel_file(unmatched_items, output_dir, filename, UNMATCHED_SHEET_TITLE, config)


# --- Private Helper Functions (Formatting and Saving Logic) ---

def _create_single_tracking_file_with_csv(
    orders: List[Dict[str, Any]],
    output_dir: str,
    run_date: datetime,
    config: Dict,
    is_consolidated: bool,
    store_id: Optional[str] = None,
) -> tuple[Optional[str], Optional[str]]:
    """
    Internal helper function to create tracking files (Excel and CSV).
    
    Returns:
        tuple: (excel_path, csv_path) - paths to generated files or None if not created
    """
    if not orders:
        return None, None

    filename = _generate_filename("Tracking", store_id, run_date, consolidated=is_consolidated, config=config)
    order_groups = _group_items_by_order_id(orders)

    rows = []
    for order_id, items in order_groups.items():
        # For the tracking file, we only need one row per order,
        # using the first item as representative.
        representative_item = items[0]
        rows.append(_format_tracking_row(representative_item, order_id))
    
    if rows:
        excel_path = _save_excel_file(rows, output_dir, filename, TRACKING_SHEET_TITLE, config, has_info_header=True)
        
        # Also create a CSV version for courier upload simulation
        csv_path = None
        if excel_path:
            csv_path = _save_tracking_csv_for_courier_upload(rows, output_dir, filename, run_date)
        
        return excel_path, csv_path
    return None, None

def _create_single_tracking_file(
    orders: List[Dict[str, Any]],
    output_dir: str,
    run_date: datetime,
    config: Dict,
    is_consolidated: bool,
    store_id: Optional[str] = None,
) -> Optional[str]:
    """
    Internal helper function to create a single tracking file (Excel only).
    Maintained for compatibility with existing code.
    """
    excel_path, _ = _create_single_tracking_file_with_csv(
        orders, output_dir, run_date, config, is_consolidated, store_id
    )
    return excel_path

def _format_run_row(item: Dict[str, Any]) -> Dict[str, Any]:
    """Formats a single item for a row in a RUN file."""
    shuffled_add = _shuffle_address(item.get('ADD1'), item.get('ADD2'), item.get('ADD3'), item.get('ADD4'))
    return {
        'FILE NAME': item.get('FILE NAME', ''),
        'Process DATE': item.get('Process DATE', ''),
        'ORIGIN OF ORDER': 'eBay',
        'FIRST NAME': item.get('FIRST NAME', ''),
        'LAST NAME': item.get('LAST NAME', ''),
        'ADD1': shuffled_add[0],
        'ADD2': shuffled_add[1],
        'ADD3': shuffled_add[2],
        'ADD4': shuffled_add[3],
        'POSTCODE': item.get('POSTCODE', ''),
        'TEL NO': item.get('TEL NO', ''),
        'EMAIL ADDRESS': item.get('EMAIL ADDRESS', ''),
        'QTY': '1',  # A RUN file is always one row per item.
        'REF NO': str(item.get('REF NO', '')).upper(),
        'TRIM': item.get('TRIM', ''),
        'Thread Colour': 'Matched',
        'Embroidery': item.get('Embroidery', ''),
        'CARPET TYPE': item.get('CARPET TYPE', ''),
        'CARPET COLOUR': item.get('CARPET COLOUR', ''),
        'Width': '',
        'Make': item.get('Make', ''),
        'Model': item.get('Model', ''),
        'YEAR': item.get('YEAR', ''),
        'Pcs/Set': item.get('Pcs/Set', ''),
        'HEEL PAD REQUIRED': 'No',
        'Other Extra': '',
        'NO OF CLIPS': item.get('NO OF CLIPS', ''),
        'CLIP TYPE': str(item.get('CLIP TYPE', '')).upper(),
        'Courier': '',
        'Tracking No': '',
        'Bar Code Type': 'CODE93',
        'Bar Code': item.get('FinalBarcode', ''), # Use the final barcode assigned by BarcodeService
        'AF': '',
        'Delivery Special Instruction': item.get('Delivery Special Instruction', ''),
        'Link to Template File': '',
        'Boot Mat 2nd SKU': '',
        'SKU': str(item.get('Raw SKU', '')).upper(),
        'Item Number': str(item.get('Item Number', '')),
        'Transaction ID': str(item.get('Transaction ID', '')),
        'ORDER ID': item.get('ORDER ID', ''),
    }

def _format_courier_master_row(first_item: Dict, all_items_in_order: List) -> Dict:
    """Formats the row for the COURIER_MASTER file."""
    shuffled_add = _shuffle_address(first_item.get('ADD1'), first_item.get('ADD2'), first_item.get('ADD3'), first_item.get('ADD4'))
    postcode = str(first_item.get('POSTCODE', '')).strip().upper()
    
    # Determine service based on postcode.
    is_highlands = any(postcode.startswith(prefix) for prefix in HIGHLANDS_AND_ISLANDS_POSTCODES)
    service = STANDARD_SERVICE_NAME if is_highlands else NEXT_DAY_SERVICE_NAME

    # Determine weight (original logic)
    weight = 15
    if len(all_items_in_order) == 1:
        item = all_items_in_order[0]
        is_ct65_black = item.get('CARPET TYPE') == 'CT65' and item.get('CARPET COLOUR', '').upper() == 'BLACK'
        weight = 1 if is_ct65_black else 15

    return {
        'Name': first_item.get('FIRST NAME', ''),
        'SURNAME': first_item.get('LAST NAME', '') or '..',
        'Address_line_1': shuffled_add[0],
        'Address_line_2': shuffled_add[1],
        'Address_line_3': shuffled_add[2],
        'Postcode': postcode,
        'BarCode': first_item.get('FinalBarcode', ''), # Use the barcode from the first item as representative of the order.
        'COUNTRY': shuffled_add[3] if len(shuffled_add[3]) <= 3 else 'GB',
        'SERVICE': service,
        'WEIGHT': weight,
        'DESCRIPTION': 'Car Mats',
    }

def _format_tracking_row(item: Dict, order_id: str) -> Dict:
    """Formats a row for a tracking file."""
    return {
        'Shipping Status': 'Shipped',
        'Order ID': str(order_id),
        'Item Number': str(item.get('Item Number', '')),
        'Item Title': item.get('Product Title', ''),
        'Custom Label': str(item.get('Raw SKU', '')).upper(),
        'Transaction ID': str(item.get('Transaction ID', '')),
        'Shipping Carrier Used': 'Hermes',
        'Tracking Number': '',
        'Barcode': str(order_id), # The "Barcode" in this file is the eBay Order ID
        'Our_Barcode': item.get('FinalBarcode', ''), # Our unique internal barcode
    }

def _generate_filename(
    file_type: str, store_id: Optional[str], current_date: datetime, extension="xlsx", consolidated=False, config=None
) -> str:
    """Standardized filename generator."""
    date_str = current_date.strftime("%Y%m%d_%H%M")
    if consolidated:
        return f"{file_type}_CONSOLIDATED_{date_str}.{extension}"
    
    store_initials_map = config.get('STORE_INITIALS', {})
    store_initial = store_initials_map.get(store_id, store_id[:3].upper() if store_id else "UNK")
    return f"{file_type}_{store_initial}_{date_str}.{extension}"

def _group_items_by_order_id(items: List[Dict]) -> Dict[str, List[Dict]]:
    """Groups a list of items into a dictionary by their Order ID."""
    groups = {}
    for item in items:
        order_id = item.get('ORDER ID')
        if order_id:
            groups.setdefault(order_id, []).append(item)
    return groups

def _group_items_by_store_id(items: List[Dict]) -> Dict[str, List[Dict]]:
    """Groups a list of items into a dictionary by their Store ID."""
    groups = {}
    for item in items:
        store_id = item.get('Store ID')
        if store_id:
            groups.setdefault(store_id, []).append(item)
    return groups

def _shuffle_address(add1, add2, add3, add4) -> List[str]:
    """Moves address parts to fill gaps, ensuring 4 parts."""
    parts = [p for p in [add1, add2, add3, add4] if p and str(p).strip().lower() not in ('', 'n/a')]
    return (parts + [''] * 4)[:4]

def _save_tracking_csv_for_courier_upload(rows: List[Dict], output_dir: str, excel_filename: str, run_date: datetime) -> Optional[str]:
    """
    Creates a CSV file that simulates courier tracking data for demo purposes.
    This allows users to test the tracking upload functionality.
    
    Returns:
        Path to the created CSV file, or None if no file was created.
    """
    logger.info(f"Starting CSV generation for {len(rows)} tracking rows")
    try:
        # Create a simplified CSV with courier format: Order Number, Consignment Number
        courier_data = []
        for i, row in enumerate(rows):
            # Use Our_Barcode as Order Number and generate a fake tracking number
            order_number = row.get('Our_Barcode', '')
            logger.debug(f"Row {i}: Our_Barcode = '{order_number}'")
            if order_number:
                # Generate a realistic looking tracking number
                tracking_number = f"HM{run_date.strftime('%y%m%d')}{str(hash(order_number))[-6:]}".upper()
                courier_data.append({
                    'Order Number': order_number,
                    'Consignment Number': tracking_number
                })
        
        logger.info(f"Generated {len(courier_data)} courier data entries")
        
        if courier_data:
            # Create CSV filename based on Excel filename
            csv_filename = excel_filename.replace('.xlsx', '_COURIER_UPLOAD_DEMO.csv')
            csv_path = os.path.join(output_dir, csv_filename)
            
            df = pd.DataFrame(courier_data)
            df.to_csv(csv_path, index=False)
            
            logger.info(f"Created courier upload demo CSV: {csv_filename}")
            return csv_path
        else:
            logger.warning("No courier data generated - no valid Our_Barcode values found")
            return None
    except Exception as e:
        logger.error(f"Failed to create courier CSV demo file: {e}", exc_info=True)
        return None

def _save_excel_file(
    rows: List[Dict[str, Any]],
    output_dir: str,
    filename: str,
    sheet_title: str,
    config: Dict,
    has_info_header: bool = False,
) -> Optional[str]:
    """
    Centralized and robust helper function to save a list of dictionaries to an Excel file.
    Uses a temporary file for atomic and safe writing.
    """
    if not rows:
        logger.warning(f"No rows to write to file {filename}. File creation will be skipped.")
        return None

    file_path = os.path.join(output_dir, filename)
    temp_path = file_path + f".{os.getpid()}.tmp"

    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_title[:30]  # Sheet title with character limit

        # Write optional #INFO header
        start_row = 1
        if has_info_header:
            ws.cell(row=1, column=1).value = INFO_HEADER_TAG
            start_row = 2
        
        # Write column headers
        headers = list(rows[0].keys())
        for col_idx, header_text in enumerate(headers, 1):
            ws.cell(row=start_row, column=col_idx).value = header_text

        # Columns that should be treated as text to avoid Excel auto-formatting
        text_format_columns = {'Barcode', 'Our_Barcode', 'ORDER ID', 'Item Number', 'Transaction ID', 'POSTCODE', 'TEL NO', 'SKU', 'Bar Code', 'Tracking No'}

        # Write data rows
        for row_idx, data_row in enumerate(rows, start_row + 1):
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                # Sanitize the value before writing it
                value = sanitize_for_excel(data_row.get(header, ''))
                cell.value = value
                # Apply text formatting if it's a critical column
                if header in text_format_columns:
                    cell.number_format = '@'

        # Adjust column widths
        for col_idx, header in enumerate(headers, 1):
            column_letter = get_column_letter(col_idx)
            # Calculate maximum width based on content and header
            max_len = max([len(str(r.get(header, ''))) for r in rows] + [len(header)])
            adjusted_width = min(max(max_len + 2, 12), 50) # Width between 12 and 50 characters
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(temp_path)
        shutil.move(temp_path, file_path)

        logger.info(f"Excel file generated successfully: {file_path}")
        return file_path

    except Exception as e:
        logger.error(f"Failed to save Excel file {filename}: {e}", exc_info=True)
        # Clean up temporary file if it exists
        if os.path.exists(temp_path):
            try:
                os.remove(temp_path)
            except OSError:
                pass
        raise FileGenerationError(f"Could not generate file {filename}", filename=filename) from e
    finally:
        if 'wb' in locals():
            wb.close()